import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time
import pandas as pd


PAYLOAD = {
    "pageMode": "Live_Site",
    "chapterId": "R8NtNFIn2t/uQ862UAXnVg==",
    "languageLocaleCode": "en_IN",
    "website_type": "1",
    "website_id": "13451",
    "mappedWidgetSettings": '[{"key":83,"name":"Chapter Website","value":"Chapter Website"},{"key":84,"name":"View Chapter Gallery","value":"View Chapter Gallery"},{"key":85,"name":"Visit This Chapter","value":"Visit This Chapter"},{"key":86,"name":"Member Names","value":"Member Name"},{"key":87,"name":"Company","value":"Company"},{"key":88,"name":"Profession/Speciality","value":"Profession/Speciality"},{"key":89,"name":"Phone","value":"Phone"},{"key":90,"name":"Send Mail","value":"Send Mail"},{"key":91,"name":"Showing","value":"Showing"},{"key":92,"name":"to","value":"to"},{"key":93,"name":"of","value":"of"},{"key":94,"name":"entries","value":"entries"},{"key":95,"name":"Meeting Details","value":"Meeting Details"},{"key":96,"name":"View Map","value":"View Map"},{"key":97,"name":"Member Count","value":"Member Count"},{"key":98,"name":"Show Members","value":"Show Members"},{"key":99,"name":"Chapter Leadership","value":"Chapter Leadership"},{"key":233,"name":"Directions","value":"Directions"},{"key":307,"name":"Zero Records","value":"Zero Records"},{"key":390,"name":"Apply Now"}]',
    "planyourvisit": "y"
}

# ==========================================================
# CONFIGURATION
# ==========================================================

AJAX_URL = "https://bni-india.in/bnicms/v3/frontend/chapterdetail/display"

HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "X-Requested-With": "XMLHttpRequest",
    "Referer": "https://bni-india.in/en-IN/chapterdetail"
}

session = requests.Session()
session.headers.update(HEADERS)




# ==========================================================
# DOWNLOAD MEMBER TABLE
# ==========================================================

print("Downloading member list...")

response = session.post(
    AJAX_URL,
    data=PAYLOAD
)

response.raise_for_status()

with open("response.html", "w", encoding="utf-8") as f:
    f.write(response.text)

print("response.html saved")

# ==========================================================
# PARSE TABLE
# ==========================================================

soup = BeautifulSoup(response.text, "html.parser")

table = soup.find("table")

if table is None:
    raise Exception("Member table not found.")

rows = table.find_all("tr")

members = []

print(f"Rows Found : {len(rows)-1}")

for row in rows[1:]:

    cols = row.find_all("td")

    if len(cols) < 4:
        continue

    name = cols[0].get_text(strip=True)

    company = cols[1].get_text(strip=True)

    profession = cols[2].get_text(strip=True)

    phone = cols[3].get_text(strip=True)

    profile = cols[0].find("a")

    from urllib.parse import urlparse, parse_qs

    profile = cols[0].find("a")

    profile_url = ""
    member_id = ""

    if profile:

        profile_url = "https://bni-india.in/en-IN/" + profile["href"]

    query = urlparse(profile["href"]).query
    params = parse_qs(query)

    member_id = params.get("encryptedMemberId", [""])[0]

    members.append({
    "Name": name,
    "Company": company,
    "Profession": profession,
    "Phone": phone,
    "Website": "",
    "Profile URL": profile_url,
    "Member ID": member_id
})

# ==========================================================
# CREATE DATAFRAME
# ==========================================================

df = pd.DataFrame(members)

options = webdriver.ChromeOptions()

# Comment this if you want to see Chrome
# options.add_argument("--headless=new")

driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()),
    options=options
)

df["Address"] = ""
df["Chapter"] = ""
# ============================================
# SCRAPE WEBSITE FROM LEADERSHIP CARD
# ============================================

print("\nFinding company websites...")



chapter_soup = BeautifulSoup(response.text, "html.parser")

cards = chapter_soup.select("div.leadership_card_content_holder")

print("Cards Found :", len(cards))

website_map = {}

for card in cards:

    company_tag = card.select_one("p.company_name a")

    if not company_tag:
        continue

    company = company_tag.get_text(strip=True)

    website = company_tag.get("href", "")

    website_map[company] = website

    print("\nWebsite Map")

for k, v in website_map.items():
    print(f"'{k}' --> {v}")


print("\nCompanies in DataFrame")
for company in df["Company"]:
    print(f"'{company}'")

for index in df.index:

    company = df.loc[index, "Company"].strip().lower()

    for key, website in website_map.items():

        if company == key.strip().lower():
            df.loc[index, "Website"] = website
            break

print("\nWebsites Found")

for company, website in website_map.items():
    print(company, "->", website)

# ============================================
# UPDATE WEBSITE COLUMN
# ============================================

    print("\nFetching Address & Chapter...")

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

for index in df.index:

    url = df.loc[index, "Profile URL"]

    print(f"{index+1}/{len(df)} -> {df.loc[index, 'Name']}")

    driver.get(url)

    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.TAG_NAME, "body"))
    )

    member_soup = BeautifulSoup(driver.page_source, "html.parser")

    # Save first page for debugging only
    if index == 0:
        with open("member_page.html", "w", encoding="utf-8") as f:
            f.write(driver.page_source)
        print("Saved member_page.html")

    # ---------------- Address ----------------
    address = ""

    address_tag = member_soup.select_one(
        "section.widgetMemberCompanyDetail h6"
    )

    if address_tag:
        address = address_tag.get_text(
            separator=", ",
            strip=True
        )

    # ---------------- Chapter ----------------
    chapter = ""

    chapter_tag = member_soup.select_one(
        ".memberContactDetails p a"
    )

    if chapter_tag:
        chapter = chapter_tag.get_text(strip=True)

    # Save to DataFrame
    df.loc[index, "Address"] = address
    df.loc[index, "Chapter"] = chapter

    print("Chapter :", chapter)
    print("Address :", address)
    print("-" * 60)

driver.quit()

print("\nFirst Five Records\n")

print(df[["Name", "Chapter", "Address"]].head())

print("\nTotal Members :", len(df))

# ==========================================================
# SAVE EXCEL
# ==========================================================
df["Website"] = df["Website"].apply(
    lambda x: "N/A" if pd.isna(x) or str(x).strip() == "" else x
)
# Save DataFrame first
df.to_excel("BNI_Members_Data.xlsx", index=False)

# Open workbook
wb = load_workbook("BNI_Members_Data.xlsx")
ws = wb.active


from openpyxl.styles import Font

ws.freeze_panes = "A2"

for cell in ws[1]:
    cell.font = Font(bold=True)

# Auto-adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter

    for cell in column:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))

    ws.column_dimensions[column_letter].width = min(max_length + 3, 60)

# Save formatted workbook
wb.save("BNI_Members_Data.xlsx")



print("\nExcel Saved Successfully")
print(f"Total Members Scraped : {len(df)}")
print("Output File : output.xlsx")




